#conding = utf-8
#/bin/env/python
import argparse
import os
import re

import openpyxl
from Bio import Entrez, SeqIO


class find_large(object):
    '''output csv(blastp) & faa(sequence)'''
    def __init__(self):
        '''input xlsx(blastp) & gbk(sequence), using gbk name'''
        parser = argparse.ArgumentParser(description='This script is used to find protein to build up a phylogenetic tree.')
        parser.add_argument('-i','--input',required=True,help='The input dir should have .gbk(origin) and .xlsx(blastp _filt.xlsx)')
        parser.add_argument('-o','--output',default='./output',required=False,help='The output dir')
        parser.add_argument('-p','--product',default='terminase large',required=False,help='The product name, defult is "terminase large"')
        parser.add_argument('-g','--gbk',required=False,help='The input gbk. This parameter is to replace the --input')
        parser.add_argument('-x','--xlsx',required=False,help='The input xlsx. This parameter is to replace the --input')
        args = parser.parse_args()
        self.main(args)

    def mkdir(self,path):
        '''make sure the path for squirrelling have no bug'''
        if not os.path.exists(path):
            os.makedirs(path)
        return path

    def input_file(self,gate):
        name_path = []
        for root, dirs, files in os.walk(gate):
            for f in files:
                if f.endswith('.gbk'):
                    name = f[:-4]
                    p = os.path.join(root,f) 
                    name_path.append([name,p])
        return name_path

    def out_fa(self,name,seq,out_path):
        with open('{}/{}.fasta'.format(out_path,name),'w') as g:
            g.write('>{} {}\n'.format(name,self.product))
            g.write(seq)

    def find_prot(self,product,gbk_path):
        '''output the protein sequence and local_number in gbk'''
        def get_cds(loc):
            start = str(int(re.findall('\[(\d+)\:(\d+)\]',str(loc))[0][0]) + 1)
            end = re.findall('\[(\d+)\:(\d+)\]',str(loc))[0][1]
            if re.findall('\((.)\)',str(loc))[0] == '-':
                cds = 'complement({0}..{1})'.format(start,end)
            elif re.findall('\((.)\)',str(loc))[0] == '+':
                cds = '{0}..{1}'.format(start,end)
            return cds
        cds_sequence = []
        records = list(SeqIO.parse(gbk_path, "genbank"))[0]
        for i in range(1,len(records)):
            if 'CDS' != records.features[i].type: # only run the protein
                continue
            try:
                if product in records.features[i].qualifiers['product']:
                    seq = records.features[i].qualifiers['translation']
                    loc = records.features[i].location
                    cds = get_cds(loc)
                    cds_sequence.append([cds,seq])
                    return cds_sequence
            except KeyError:
                continue
            except IndexError:
                print('Total features: {}'.format(i-1))
                exit('Python could not find "{}" in your gbk file {}.  QAQ'.format(product,gbk_path))

    def find_xlsx(self,name,loc_num,input_dir):
        '''ouput csv, using local_number find protein blastp infomation in xlsx'''
        try:
            wb = openpyxl.load_workbook('{}/{}.xlsx'.format(input_dir,name))
            sheet = wb[loc_num]
            flag = 0 # old is the ncbi download
            return sheet,flag
        except FileNotFoundError:
            wb = openpyxl.load_workbook('{}/{}_filt.xlsx'.format(input_dir,name))
            sheet = wb[loc_num]
            flag = 1 # new is the entrez parsed
            return sheet,flag

    def id_filter(self,sh,flag):
        ncbi_id = []
        i = 1
        if flag == 0:
            while True:
                try:
                    i += 1
                    if float(sh.cell(row=i, column=9).value) > 1e-3:
                        continue
                    elif float(sh.cell(row=i, column=8).value.strip('%')) < 50:
                        continue
                    elif float(sh.cell(row=i, column=10).value.strip('%')) < 50:
                        continue
                    else:
                        id = sh.cell(row=i, column=12).value
                        if not id in ncbi_id:
                            ncbi_id.append(id)
                except TypeError:
                    break
        else:
            while True:
                try:
                    i += 1
                    if float(sh.cell(row=i, column=5).value) > 1e-3:
                        continue
                    elif float(sh.cell(row=i, column=4).value) < 50:
                        continue
                    elif float(sh.cell(row=i, column=3).value) < 50:
                        continue
                    else:
                        id = sh.cell(row=i, column=6).value
                        if not id in ncbi_id:
                            ncbi_id.append(id)
                except TypeError:
                    break
        return ncbi_id
    
    def to_download(self,out_path,ncbi_id,re_times):
        try:
            if not os.path.exists(out_path):
                with open(out_path,'w') as g:
                    efetch_seq = Entrez.efetch(db='protein', id=ncbi_id, rettype='fasta', retmode='text')
                    g.write(efetch_seq.read())
                print('...finished, times of retrying: ' + str(re_times))
        except TimeoutError:
            self.to_download(out_path,ncbi_id,re_times)
        except ConnectionError:
            self.to_download(out_path,ncbi_id,re_times)
        re_times += 1
    
    def extracting(self,ncbi_ids,output_path):
        for ncbi_id in ncbi_ids:
            print('downloading: {}'.format(ncbi_id),end='')
            re_times = 0
            out_path = os.path.join(output_path,'{}.faa'.format(ncbi_id))
            self.to_download(out_path,ncbi_id,re_times)

    def format_name(self,ide):
        rstr = r"[\/\\\:\*\?\"\<\>\|]"
        new_ide = re.sub(rstr, "_", ide)
        return new_ide

    def rename(self,gate):
        '''rename the NCBI download .fasta with its fasta_title'''
        for i in os.listdir(gate):
            try:
                path = gate + '/' + i
                if i.endswith('.faa'):
                    with open(path,'r') as f:
                        title = f.readline().replace('\n','')
                    name = self.format_name(title.split(' ')[-1].replace(']','').replace('[','') + '.faa')
                    if not os.path.exists(gate + '/' + name):
                        os.rename(path,gate + '/' + name)
            except:
                print('Error: rename')
                continue
    
    def main(self,args):
        out_path = args.output
        product = args.product
        self.mkdir(out_path)
        input_dir = args.input
        ncbi_ids = []
        for name,path in self.input_file(input_dir):
            print('\ncurrent:{}\n'.format(name))
            local_number,seq = self.find_prot(product,path)
            self.out_fasta = self.out_fa(name,seq,out_path)
            sheet,flag = self.find_xlsx(name,local_number,input_dir)
            ncbi_ids = self.id_filter(sheet,flag)
            self.extracting(ncbi_ids,out_path)
            self.rename(out_path)

if __name__ == '__main__':
    find_large()
